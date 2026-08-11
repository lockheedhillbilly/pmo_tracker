"""Web dashboard for the PMO tracker (deploys to Vercel via dashboard:app,
see pyproject.toml) and reads/writes the same database the chat tools and
email scripts use — one source of truth.

Covers: pivoting (view by / then by), inline + group-level add rows with
paste-to-parse, a dedicated Capture tab for pasting notes/emails, a
fixed/sticky completion column, Excel-style filters with active-filter
chips, bulk actions, row + column drag reorder, a column manager, saved
views, a review workflow, per-task notes, an audit trail, task
dependencies, custom fields, a formatted Excel export, a day-by-day
activity browser, a project-context doc editable from its own tab, and
keyboard shortcuts. See chat for what was deliberately simplified or
deferred (note @mentions/link previews, a "created by" column on tasks
themselves — there's no identity system across chat/email/dashboard to
support one truthfully, so history entries fall back to "Unknown" for
anything not made through this dashboard).
"""

import io
import os
import sys
from pathlib import Path

from flask import Flask, Response, jsonify, render_template, request
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

sys.path.insert(0, str(Path(__file__).parent))
from db import (  # noqa: E402
    EXECUTION_STATES, PRIORITIES, REVIEW_STATUSES, REVIEW_TYPES, STATUSES, TRACKS,
    TaskStore, TrackerError,
)
import email_recipients  # noqa: E402
import email_send  # noqa: E402
import excel_export  # noqa: E402
import excel_import  # noqa: E402
import nlu  # noqa: E402

DEFAULT_DB_PATH = Path(__file__).parent / "tasks.db"
DB_PATH = os.environ.get("PMO_TRACKER_DB_PATH", str(DEFAULT_DB_PATH))
CURRENT_USER = "Akshit"  # the one and only reviewer; always offered as an owner too
store = TaskStore(DB_PATH)

app = Flask(__name__)


@app.get("/")
def index():
    resp = Response(render_template("index.html", current_user=CURRENT_USER), mimetype="text/html")
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate"
    return resp


@app.get("/api/meta")
def meta():
    tasks = store.list_tasks()
    modules = sorted({t["module"] for t in tasks if t["module"]})
    collaborator_names = {
        n.strip() for t in tasks for n in (t["collaborators"] or "").split(",") if n.strip()
    }
    owners = sorted({t["owner"] for t in tasks} | collaborator_names | {CURRENT_USER})
    return jsonify({
        "tracks": list(TRACKS), "priorities": list(PRIORITIES), "statuses": list(STATUSES),
        "execution_states": list(EXECUTION_STATES), "review_statuses": list(REVIEW_STATUSES),
        "review_types": list(REVIEW_TYPES), "modules": modules, "owners": owners,
    })


@app.get("/api/tasks")
def get_tasks():
    return jsonify(store.list_tasks())


@app.get("/api/note_counts")
def note_counts():
    return jsonify(store.note_summaries())


@app.post("/api/tasks")
def create_task():
    try:
        return jsonify(store.add_task(**request.get_json(force=True)))
    except TrackerError as e:
        return jsonify({"error": str(e)}), 400


@app.patch("/api/tasks/<int:task_id>")
def edit_task(task_id):
    body = request.get_json(force=True)
    body.setdefault("changed_by", CURRENT_USER)
    try:
        return jsonify(store.update_task(id=task_id, **body))
    except TrackerError as e:
        return jsonify({"error": str(e)}), 400


@app.get("/api/tasks/<int:task_id>/history")
def get_history(task_id):
    return jsonify(store.list_history(task_id))


@app.get("/api/history")
def get_recent_history():
    return jsonify(store.list_recent_history(
        since=request.args.get("since"), until=request.args.get("until"),
    ))


@app.get("/api/export.xlsx")
def export_xlsx():
    tasks = store.list_tasks()
    headers = [
        "Track", "Use case", "Owner", "Collaborators", "Task", "Added", "Due", "Priority",
        "Status", "Execution", "Review status", "Blocked by",
    ]
    by_id = {t["id"]: t["task"] for t in tasks}

    wb = Workbook()
    ws = wb.active
    ws.title = "PMO Tasks"
    ws.append(headers)
    for t in tasks:
        ws.append([
            t["track"], t["module"] or "", t["owner"], t["collaborators"] or "", t["task"],
            t["added"], t["due"], t["priority"], t["status"], t["execution_state"] or "",
            t["review_status"] or "", by_id.get(t["blocked_by_id"], ""),
        ])

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1B4D3E", end_color="1B4D3E", fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 20

    widths = [14, 20, 14, 20, 48, 12, 12, 10, 10, 14, 16, 32]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # An Excel Table (not just auto_filter) gives filter dropdowns AND banded
    # row styling in one step — this is the "pre-loaded filters" part.
    last_col = get_column_letter(len(headers))
    last_row = len(tasks) + 1
    if len(tasks) > 0:
        table = Table(displayName="PMOTasks", ref=f"A1:{last_col}{last_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False,
        )
        ws.add_table(table)
    else:
        ws.auto_filter.ref = f"A1:{last_col}1"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = Response(
        buf.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp.headers["Content-Disposition"] = "attachment; filename=pmo_tasks.xlsx"
    return resp


@app.get("/api/gantt/export.xlsx")
def export_gantt_xlsx():
    data = excel_export.build_workbook(store.list_tasks())
    resp = Response(data, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp.headers["Content-Disposition"] = "attachment; filename=pmo_gantt.xlsx"
    return resp


@app.post("/api/gantt/import/preview")
def preview_gantt_import():
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "No file uploaded"}), 400
    try:
        return jsonify(excel_import.parse_file(f.filename, f.read()))
    except Exception as e:  # noqa: BLE001 — surface any parse failure (bad/corrupt file) as a 400
        return jsonify({"error": str(e)}), 400


@app.post("/api/gantt/import/commit")
def commit_gantt_import():
    body = request.get_json(force=True)
    rows, mapping = body.get("rows", []), body.get("mapping", {})
    return jsonify(excel_import.commit_rows(store, rows, mapping, changed_by=CURRENT_USER))


@app.get("/api/email/recipients")
def get_email_recipients():
    return jsonify(email_recipients.list_recipients(store))


@app.post("/api/email/recipients")
def add_email_recipient():
    body = request.get_json(force=True)
    try:
        return jsonify(email_recipients.add_recipient(store, body.get("name", ""), body.get("email", "")))
    except ValueError as e:
        return jsonify({"error": str(e)}), 400


@app.delete("/api/email/recipients/<path:email_addr>")
def delete_email_recipient(email_addr):
    return jsonify(email_recipients.remove_recipient(store, email_addr))


@app.post("/api/email/compose")
def compose_email():
    body = request.get_json(force=True)
    to_emails = body.get("to") or []
    if not to_emails:
        return jsonify({"error": "No recipients selected"}), 400
    try:
        data = excel_export.build_workbook(store.list_tasks())
        email_send.compose_with_attachment(
            to_emails,
            body.get("subject") or "PMO Tracker — latest plan",
            body.get("note") or "Attached is the latest version of the plan.",
            data, "pmo_gantt.xlsx",
        )
        return jsonify({"ok": True})
    except Exception as e:  # noqa: BLE001 — most likely cause: bad Gmail credentials
        return jsonify({"error": str(e)}), 400


@app.delete("/api/tasks/<int:task_id>")
def remove_task(task_id):
    try:
        store.delete_task(task_id)
        return jsonify({"ok": True})
    except TrackerError as e:
        return jsonify({"error": str(e)}), 400


@app.post("/api/reorder")
def reorder():
    ids = request.get_json(force=True)["ids"]
    store.reorder(ids)
    return jsonify({"ok": True})


@app.post("/api/bulk")
def bulk_update():
    body = request.get_json(force=True)
    ids, fields = body["ids"], body["fields"]
    fields.setdefault("changed_by", CURRENT_USER)
    results = []
    for task_id in ids:
        try:
            results.append(store.update_task(id=task_id, **fields))
        except TrackerError as e:
            results.append({"id": task_id, "error": str(e)})
    return jsonify(results)


@app.post("/api/parse")
def parse_text():
    body = request.get_json(force=True)
    text, defaults = body.get("text", ""), body.get("defaults", {})
    try:
        client = nlu.get_client()
        open_tasks = store.list_tasks(status="Open")
        project_context = nlu.get_project_context(store)
        result = nlu.call_claude(client, text, open_tasks, defaults=defaults, project_context=project_context)
        changes, results = nlu.apply_actions(store, result.get("actions", []), changed_by=CURRENT_USER)
        return jsonify({"changes": changes, "results": results, "notes": result.get("notes", "")})
    except Exception as e:
        return jsonify({"error": str(e)}), 400


@app.get("/api/tasks/<int:task_id>/notes")
def get_notes(task_id):
    return jsonify(store.list_notes(task_id))


@app.post("/api/tasks/<int:task_id>/notes")
def add_note(task_id):
    body = request.get_json(force=True)
    try:
        return jsonify(store.add_note(task_id, body["author"], body["text"]))
    except TrackerError as e:
        return jsonify({"error": str(e)}), 400


@app.patch("/api/notes/<int:note_id>")
def edit_note(note_id):
    body = request.get_json(force=True)
    try:
        return jsonify(store.edit_note(note_id, body["author"], body["text"]))
    except TrackerError as e:
        return jsonify({"error": str(e)}), 400


@app.delete("/api/notes/<int:note_id>")
def remove_note(note_id):
    author = request.args.get("author")
    try:
        store.delete_note(note_id, author=author)
        return jsonify({"ok": True})
    except TrackerError as e:
        return jsonify({"error": str(e)}), 400


@app.post("/api/notes/<int:note_id>/pin")
def pin_note(note_id):
    try:
        return jsonify(store.toggle_pin_note(note_id))
    except TrackerError as e:
        return jsonify({"error": str(e)}), 400


@app.get("/api/project-context")
def get_project_context():
    return jsonify({"content": nlu.get_project_context(store)})


@app.post("/api/project-context")
def save_project_context():
    body = request.get_json(force=True)
    store.set_setting("project_context", body.get("content", ""))
    return jsonify({"ok": True})


@app.get("/api/meetings")
def get_meetings():
    # A separate local process writes meeting records directly to Turso, not
    # through this API — sync here (and in run_cycle.py) picks up any
    # newly-'done' meeting's next_steps as real tasks before listing.
    store.sync_meeting_tasks()
    return jsonify(store.list_meetings())


@app.get("/api/meetings/<int:meeting_id>")
def get_meeting_detail(meeting_id):
    meeting = store.get_meeting(meeting_id)
    if meeting is None:
        return jsonify({"error": "not found"}), 404
    return jsonify(meeting)


if __name__ == "__main__":
    app.run(port=int(os.environ.get("PMO_TRACKER_PORT", 5057)), debug=False)
