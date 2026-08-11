"""Builds the Gantt Excel export workbook — Project Plan / Timeline / Milestones /
Critical Path / Risks sheets, human-readable (bold headers, frozen header row, sane
column widths) rather than a raw table dump.
"""

from __future__ import annotations

import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from schedule import compute_schedule

HEADER_FONT = Font(bold=True)


def _autosize(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_header(ws, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
    ws.freeze_panes = "A2"


def _status_label(t: dict) -> str:
    if t["status"] == "Done":
        return "Complete"
    return t.get("execution_state") or "Not started"


def build_workbook(tasks: list[dict]) -> bytes:
    by_id = {t["id"]: t for t in tasks}
    schedule = compute_schedule(tasks)
    today = date.today()

    wb = Workbook()

    ws = wb.active
    ws.title = "Project Plan"
    _write_header(ws, ["ID", "Workstream", "Task", "Owner", "Status", "Start", "End",
                        "Duration (d)", "% Complete", "Depends On", "Milestone", "Pinned"])
    for t in tasks:
        dep = by_id.get(t.get("blocked_by_id"))
        ws.append([
            t["id"], t.get("module") or "", t["task"], t["owner"], _status_label(t),
            t.get("start_date") or "", t.get("due") or "", schedule[t["id"]]["duration_days"],
            t.get("percent_complete") or 0, f"#{dep['id']} {dep['task']}" if dep else "",
            "Yes" if t.get("is_milestone") else "", "Yes" if t.get("pinned") else "",
        ])
    _autosize(ws, [6, 20, 40, 14, 12, 12, 12, 12, 12, 30, 10, 8])

    ws2 = wb.create_sheet("Timeline")
    _write_header(ws2, ["ID", "Task", "Owner", "Computed Start", "Computed Finish", "Slack (d)", "Critical"])
    for t in sorted(tasks, key=lambda t: schedule[t["id"]]["start"]):
        s = schedule[t["id"]]
        ws2.append([t["id"], t["task"], t["owner"], s["start"], s["finish"], s["slack_days"], "Yes" if s["critical"] else "No"])
    _autosize(ws2, [6, 40, 14, 14, 14, 10, 10])

    ws3 = wb.create_sheet("Milestones")
    _write_header(ws3, ["ID", "Milestone", "Workstream", "Owner", "Due", "Status"])
    milestones = [t for t in tasks if t.get("is_milestone")]
    for t in sorted(milestones, key=lambda t: t.get("due") or ""):
        ws3.append([t["id"], t["task"], t.get("module") or "", t["owner"], t.get("due") or "", _status_label(t)])
    _autosize(ws3, [6, 40, 20, 14, 12, 14])

    ws4 = wb.create_sheet("Critical Path")
    _write_header(ws4, ["ID", "Task", "Owner", "Start", "Finish", "Slack (d)"])
    critical = [t for t in tasks if t["status"] == "Open" and schedule[t["id"]]["critical"]]
    for t in sorted(critical, key=lambda t: schedule[t["id"]]["start"]):
        s = schedule[t["id"]]
        ws4.append([t["id"], t["task"], t["owner"], s["start"], s["finish"], s["slack_days"]])
    _autosize(ws4, [6, 40, 14, 14, 14, 10])

    ws5 = wb.create_sheet("Risks")
    _write_header(ws5, ["ID", "Task", "Owner", "Due", "Risk"])
    today_iso = today.isoformat()
    for t in tasks:
        if t["status"] != "Open":
            continue
        s = schedule[t["id"]]
        reasons = []
        due = t.get("due") or ""
        if due and due < today_iso:
            reasons.append("Overdue")
        if t.get("execution_state") == "Blocked":
            reasons.append("Blocked")
        if not t.get("owner"):
            reasons.append("No owner assigned")
        if s["critical"] and due and due <= today_iso:
            reasons.append("Critical path task due soon")
        if reasons:
            ws5.append([t["id"], t["task"], t["owner"], due, ", ".join(reasons)])
    _autosize(ws5, [6, 40, 14, 12, 40])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
