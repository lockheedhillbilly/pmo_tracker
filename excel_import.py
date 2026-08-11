"""Excel/CSV import for the Gantt tab. Two-step flow (spec: "show preview + column-mapping
screen before import"): parse_file() detects columns by header synonym and returns a
preview; the client shows a mapping UI, then commit_rows() actually creates tasks. Colored-
cell-only spreadsheets (no clean date columns) are out of scope — see chat.
"""

from __future__ import annotations

import csv
import io
from datetime import date, datetime

from openpyxl import load_workbook

SYNONYMS = {
    "module": ["workstream", "module", "use case", "project", "category"],
    "task": ["task", "name", "title", "activity", "task name"],
    "owner": ["owner", "assignee", "assigned to", "responsible"],
    "status": ["status", "state", "execution"],
    "start_date": ["start", "start date", "begin"],
    "due": ["end", "due", "end date", "finish", "due date", "finish date"],
    "blocked_by": ["depends on", "predecessor", "predecessors", "dependency", "dependencies", "blocked by"],
    "is_milestone": ["milestone", "is milestone"],
}


def _guess_mapping(headers: list[str]) -> dict[str, int]:
    lowered = [h.strip().lower() for h in headers]
    mapping = {}
    for field, syns in SYNONYMS.items():
        for i, h in enumerate(lowered):
            if h in syns or any(s in h for s in syns):
                mapping[field] = i
                break
    return mapping


def _cell_str(v) -> str:
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return "" if v is None else str(v)


def parse_file(filename: str, data: bytes) -> dict:
    if filename.lower().endswith(".csv"):
        text = data.decode("utf-8-sig", errors="replace")
        rows = list(csv.reader(io.StringIO(text)))
    else:
        wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = [[_cell_str(c.value) for c in row] for row in ws.iter_rows()]
    rows = [r for r in rows if any(str(c).strip() for c in r)]
    if not rows:
        return {"columns": [], "rows": [], "suggested_mapping": {}}
    headers, data_rows = rows[0], rows[1:]
    return {
        "columns": headers,
        # Cap, not a "preview vs. full" split — the client holds exactly this list and sends
        # it back unchanged on commit, so this is also the ceiling on rows actually imported.
        "rows": data_rows[:500],
        "suggested_mapping": _guess_mapping(headers),
    }


def _norm_date(s: str | None) -> str | None:
    """Only accept clean ISO dates — a fuzzy multi-format parser risks silently
    misreading day/month order, which is worse than just leaving the field blank."""
    if not s:
        return None
    try:
        date.fromisoformat(s.strip()[:10])
        return s.strip()[:10]
    except ValueError:
        return None


def commit_rows(store, rows: list[list], mapping: dict[str, int], changed_by: str) -> dict:
    """Creates one task per row, then resolves Depends On by exact task-name match within
    this same import batch — the only predecessor reference a plain spreadsheet gives us."""
    task_idx = mapping.get("task")
    if task_idx is None:
        return {"created": 0, "errors": ["No Task column mapped"]}

    def col(row, field):
        idx = mapping.get(field)
        if idx is None or idx >= len(row):
            return None
        val = row[idx]
        val = val.strip() if isinstance(val, str) else val
        return val or None

    created_by_name: dict[str, int] = {}
    created_ids: list[int] = []
    pending_deps: list[tuple[int, str]] = []
    errors: list[str] = []

    for row_num, row in enumerate(rows, start=2):
        task_name = col(row, "task")
        if not task_name:
            continue
        status_raw = (col(row, "status") or "").strip().lower()
        status = "Done" if status_raw in ("complete", "done", "closed") else "Open"
        execution_state = status_raw.capitalize() if status_raw in ("not started", "in progress", "blocked") else None
        milestone_raw = (col(row, "is_milestone") or "").strip().lower()
        try:
            created = store.add_task(
                track="Discovery", owner=col(row, "owner") or changed_by, task=str(task_name),
                module=col(row, "module"), due=_norm_date(col(row, "due")), status=status,
                execution_state=execution_state, start_date=_norm_date(col(row, "start_date")),
                is_milestone=milestone_raw in ("yes", "true", "1", "milestone"),
            )
        except Exception as e:  # noqa: BLE001 — one bad row shouldn't abort the whole import
            errors.append(f"Row {row_num} ({task_name}): {e}")
            continue
        created_ids.append(created["id"])
        created_by_name[str(task_name).strip().lower()] = created["id"]
        dep_name = col(row, "blocked_by")
        if dep_name:
            pending_deps.append((created["id"], str(dep_name).strip().lower()))

    for task_id, dep_name in pending_deps:
        dep_id = created_by_name.get(dep_name)
        if dep_id:
            try:
                store.update_task(id=task_id, blocked_by_id=dep_id, changed_by=changed_by)
            except Exception as e:  # noqa: BLE001
                errors.append(f"Could not link dependency for task {task_id}: {e}")

    return {"created": len(created_ids), "errors": errors}
